from typing import Any, List, Dict
import openpyxl
from tkinter import messagebox

def parse_athlete_data(athlete_div: Any, network_manager) -> Dict[str, Any]:
    nome = athlete_div.find(class_='card-title').text
    età = int(athlete_div.find(class_='card-title').find_next_sibling(class_='card-title').text.split(':')[-1])
    società = athlete_div.find('h6', string='Società').find_next('p').text

    bottone = athlete_div.find('button', class_='btn btn-dark btn-sm record')
    matricola = bottone["data-id"]
    statistiche = network_manager.get_athlete_stats(matricola)

    return {
        "nome": nome,
        "età": età,
        "società": società,
        "statistiche": statistiche
    }

def filter_athletes(athletes: List[Dict[str, Any]], min_matches: int, max_matches: int) -> List[Dict[str, Any]]:
    return [
        atleta for atleta in athletes
        if min_matches <= atleta["statistiche"]["numero_match"] <= max_matches
    ]

def save_to_excel(athletes: List[Dict[str, Any]], file_name: str) -> None:
    wb = openpyxl.Workbook()
    sheet = wb.active

    keys = ["Nome e cognome", "Età", "Società", "Numero match", "Vittorie", "Sconfitte", "Pareggi"]
    for col_num, key in enumerate(keys, start=1):
        sheet.cell(row=1, column=col_num, value=key)

    for row_num, atleta in enumerate(athletes, start=2):
        sheet.cell(row=row_num, column=1, value=atleta.get("nome"))
        sheet.cell(row=row_num, column=2, value=atleta.get("età"))
        sheet.cell(row=row_num, column=3, value=atleta.get("società"))
        sheet.cell(row=row_num, column=4, value=atleta["statistiche"].get("numero_match"))
        sheet.cell(row=row_num, column=5, value=atleta["statistiche"].get("vittorie"))
        sheet.cell(row=row_num, column=6, value=atleta["statistiche"].get("sconfitte"))
        sheet.cell(row=row_num, column=7, value=atleta["statistiche"].get("pareggi"))

    try:
        wb.save(f"{file_name}.xlsx")
        messagebox.showinfo("Successo", f"File '{file_name}.xlsx' creato con successo!")
    except Exception as e:
        messagebox.showerror("Errore", f"Errore durante la creazione del file Excel:\n{str(e)}")
